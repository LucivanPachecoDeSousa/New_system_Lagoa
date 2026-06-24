from PySide6.QtWidgets import QFileDialog, QMessageBox
from utils.widgets import msg_box


def exportar_excel(parent, nome_sugerido, titulo_aba, cabecalhos, dados, col_decimals=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        msg_box(parent, QMessageBox.Critical, "Erro", f"openpyxl não encontrado: {e}")
        return

    path, _ = QFileDialog.getSaveFileName(
        parent, "Exportar Relatório", nome_sugerido, "Excel (*.xlsx)"
    )
    if not path:
        return

    try:
        wb = Workbook()
        ws = wb.active
        import re
        sanitized = re.sub(r'[\\/*?:\[\]]', '', titulo_aba)[:31]
        ws.title = sanitized if sanitized else "Relatorio"

        cols = len(cabecalhos)

        header_fill = PatternFill(start_color="795548", end_color="795548", fill_type="solid")
        header_font = Font(color="ffffff", bold=True, size=11, name="Calibri")
        thin_border = Border(
            left=Side(style='thin', color="cccccc"),
            right=Side(style='thin', color="cccccc"),
            top=Side(style='thin', color="cccccc"),
            bottom=Side(style='thin', color="cccccc"),
        )

        ws.row_dimensions[1].height = 28
        for col, titulo in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col, value=titulo)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for i, linha in enumerate(dados, 2):
            ws.row_dimensions[i].height = 20
            for col, val in enumerate(linha, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                cell.font = Font(size=10, name="Calibri")
                if isinstance(val, float):
                    dec = (col_decimals or {}).get(col, 2)
                    cell.number_format = f'#,##0.{"0" * dec}'
                    cell.alignment = Alignment(horizontal='right')
                elif isinstance(val, int):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.column_dimensions['A'].width = 6
        for col in range(2, cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.tabColor = "795548"

        wb.save(path)

        msg_box(parent, QMessageBox.Information, "Exportado", f"Relatório salvo em:\n{path}")
    except Exception as e:
        msg_box(parent, QMessageBox.Critical, "Erro ao Exportar",
                f"Não foi possível salvar o relatório:\n{e}")
