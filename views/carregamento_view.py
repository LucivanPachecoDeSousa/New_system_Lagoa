from datetime import date
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QComboBox, QFrame, QDialog, QDateEdit, QInputDialog,
    QAbstractItemView, QGraphicsDropShadowEffect, QDoubleSpinBox,
    QFileDialog,
)
from PySide6.QtCore import Qt, QDate, QTimer
from PySide6.QtGui import QColor
from controllers.carregamento_controller import CarregamentoController
from controllers.auth_controller import AuthController
from utils.widgets import estilizar_calendario, msg_box


class CarregamentoDialog(QDialog):
    def __init__(self, parent=None, carregamento=None):
        super().__init__(parent)
        self.carregamento = carregamento
        self.controller = CarregamentoController()
        self.setWindowTitle("Editar Carregamento" if carregamento else "Novo Carregamento")
        self.setMinimumWidth(950)
        self.setMinimumHeight(750)
        self.setModal(True)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self._setup_ui()
        self.showMaximized()
        if carregamento:
            self._preencher(carregamento)

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)

        card = QFrame()
        card.setMinimumSize(900, 700)
        card.setStyleSheet("QFrame { background: white; border-radius: 20px; }")

        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(60)
        shadow.setColor(QColor(0, 0, 0, 40))
        shadow.setOffset(0, 10)
        card.setGraphicsEffect(shadow)

        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(30, 20, 30, 20)
        card_layout.setSpacing(8)

        title = QLabel("Editar Carregamento" if self.carregamento else "Novo Carregamento")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #4E342E; font-size: 20px; font-weight: 700; letter-spacing: 2px;")
        card_layout.addWidget(title)

        self._add_field_row(card_layout, "DATA", self._create_date())
        self._add_field_row(card_layout, "ENTIDADE", self._create_entidade())
        self._add_field_row(card_layout, "LOTE", self._create_lote())

        self._add_field_row(card_layout, "PLACA", self._create_placa())

        linha1 = QHBoxLayout()
        linha1.setSpacing(10)
        linha1.addWidget(self._field_widget("PESO BRUTO (KG)", self._create_peso_bruto()))
        linha1.addWidget(self._field_widget("TARA (KG)", self._create_tara()))
        linha1.addWidget(self._field_widget("PESO LÍQUIDO (KG)", self._create_peso_liquido()))
        card_layout.addLayout(linha1)

        linha2 = QHBoxLayout()
        linha2.setSpacing(10)
        linha2.addWidget(self._field_widget("PESO TICKET (KG)", self._create_peso_ticket()))
        linha2.addWidget(self._field_widget("PESO NF (KG)", self._create_peso_nf()))
        card_layout.addLayout(linha2)

        self._add_field_row(card_layout, "NÚMERO NF", self._create_numero_nf())

        linha3 = QHBoxLayout()
        linha3.setSpacing(10)
        linha3.addWidget(self._field_widget("VALOR UNITÁRIO (R$)", self._create_valor_unitario()))
        linha3.addWidget(self._field_widget("TOTAL DA NOTA (R$)", self._create_total_nota()))
        card_layout.addLayout(linha3)

        self._add_field_row(card_layout, "CHAVE NF", self._create_chave_nf())

        card_layout.addStretch()

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.setCursor(Qt.PointingHandCursor)
        btn_cancelar.setFixedHeight(44)
        btn_cancelar.setStyleSheet("""
            QPushButton {
                padding: 10px; background: #e0e0e0; color: #333;
                border: none; border-radius: 10px; font-weight: 700; font-size: 13px;
            }
            QPushButton:hover { background: #ccc; }
        """)
        btn_cancelar.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancelar)

        btn_salvar = QPushButton("Salvar")
        btn_salvar.setCursor(Qt.PointingHandCursor)
        btn_salvar.setFixedHeight(44)
        btn_salvar.setStyleSheet("""
            QPushButton {
                padding: 10px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5D4037, stop:1 #795548);
                color: white; border: none; border-radius: 10px;
                font-size: 14px; font-weight: 700; letter-spacing: 2px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #795548, stop:1 #8D6E63);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3E2723, stop:1 #5D4037);
            }
        """)
        btn_salvar.clicked.connect(self._validar_salvar)
        btn_layout.addWidget(btn_salvar)

        card_layout.addLayout(btn_layout)
        layout.addWidget(card)

        self._on_lote_changed()

    def _field_widget(self, label, widget):
        container = QWidget()
        vl = QVBoxLayout(container)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(2)
        lbl = QLabel(label)
        lbl.setStyleSheet("color: #555; font-size: 11px; font-weight: 700; letter-spacing: 1.5px;")
        vl.addWidget(lbl)
        vl.addWidget(widget)
        return container

    def _add_field_row(self, card_layout, label, widget):
        lbl = QLabel(label)
        lbl.setStyleSheet("color: #555; font-size: 11px; font-weight: 700; letter-spacing: 1.5px;")
        card_layout.addWidget(lbl)
        card_layout.addWidget(widget)

    def _input_style(self):
        return """
            QLineEdit, QComboBox, QDoubleSpinBox, QDateEdit {
                padding: 10px 14px;
                border: 2px solid #ddd;
                border-radius: 10px;
                font-size: 13px;
                background: #fafafa;
                color: #000;
            }
            QLineEdit:focus, QComboBox:focus, QDoubleSpinBox:focus, QDateEdit:focus {
                border-color: #795548;
                background: white;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 30px;
                border-left: 1px solid #ddd;
                border-top-right-radius: 10px;
                border-bottom-right-radius: 10px;
            }
            QComboBox QAbstractItemView {
                color: #000;
                background: white;
                selection-background-color: #e6e6e6;
                selection-color: #000;
            }
        """

    def _create_date(self):
        self.dt_data = QDateEdit()
        self.dt_data.setDate(QDate.currentDate())
        self.dt_data.setCalendarPopup(True)
        self.dt_data.setDisplayFormat("dd/MM/yyyy")
        estilizar_calendario(self.dt_data)
        self.dt_data.setStyleSheet(self._input_style())
        return self.dt_data

    def _create_entidade(self):
        self.cmb_entidade = QComboBox()
        self.cmb_entidade.setStyleSheet(self._input_style())
        self.cmb_entidade.addItem("Selecione a entidade", None)
        entidades = self.controller.listar_clientes_carregamento()
        for e in entidades:
            self.cmb_entidade.addItem(e["razao_social"], e["id"])
        self.cmb_entidade.currentIndexChanged.connect(self._on_entidade_changed)
        return self.cmb_entidade

    def _create_lote(self):
        self.cmb_lote = QComboBox()
        self.cmb_lote.setStyleSheet(self._input_style())
        self.cmb_lote.currentIndexChanged.connect(self._on_lote_changed)
        self._lotes = self.controller.listar_lotes_carregamento()
        self._popular_lotes()
        return self.cmb_lote

    def _popular_lotes(self, entidade_id=None):
        self.cmb_lote.blockSignals(True)
        self.cmb_lote.clear()
        self.cmb_lote.addItem("Selecione o lote", None)
        for l in self._lotes:
            if entidade_id is not None and l.get("entidade_id") == entidade_id:
                self.cmb_lote.addItem(l["nome_lote"], l["id"])
        self.cmb_lote.blockSignals(False)

    def _on_entidade_changed(self):
        entidade_id = self.cmb_entidade.currentData()
        self._popular_lotes(entidade_id)
        self._on_lote_changed()

    def _on_lote_changed(self):
        if not hasattr(self, 'spn_valor_unitario'):
            return
        valor = self.cmb_lote.currentData()
        if valor is not None:
            for l in self._lotes:
                if l["id"] == valor:
                    if l.get("valor_unitario"):
                        self.spn_valor_unitario.setValue(float(l["valor_unitario"]))
                    self._calcular_total_nota()
                    break

    def _create_placa(self):
        self.txt_placa = QLineEdit()
        self.txt_placa.setPlaceholderText("AAA-0000")
        self.txt_placa.setStyleSheet(self._input_style())
        self.txt_placa.setMaxLength(8)
        return self.txt_placa

    def _create_peso_bruto(self):
        self.spn_peso_bruto = QDoubleSpinBox()
        self.spn_peso_bruto.setRange(0, 999999)
        self.spn_peso_bruto.setDecimals(0)
        self.spn_peso_bruto.setStyleSheet(self._input_style())
        self.spn_peso_bruto.valueChanged.connect(self._calcular_peso_liquido)
        return self.spn_peso_bruto

    def _create_tara(self):
        self.spn_tara = QDoubleSpinBox()
        self.spn_tara.setRange(0, 999999)
        self.spn_tara.setDecimals(0)
        self.spn_tara.setStyleSheet(self._input_style())
        self.spn_tara.valueChanged.connect(self._calcular_peso_liquido)
        return self.spn_tara

    def _create_peso_liquido(self):
        self.spn_peso_liquido = QDoubleSpinBox()
        self.spn_peso_liquido.setRange(0, 999999)
        self.spn_peso_liquido.setDecimals(0)
        self.spn_peso_liquido.setStyleSheet(self._input_style())
        self.spn_peso_liquido.setReadOnly(True)
        return self.spn_peso_liquido

    def _create_peso_ticket(self):
        self.spn_peso_ticket = QDoubleSpinBox()
        self.spn_peso_ticket.setRange(0, 999999)
        self.spn_peso_ticket.setDecimals(0)
        self.spn_peso_ticket.setStyleSheet(self._input_style())
        return self.spn_peso_ticket

    def _create_peso_nf(self):
        self.spn_peso_nf = QDoubleSpinBox()
        self.spn_peso_nf.setRange(0, 999999)
        self.spn_peso_nf.setDecimals(0)
        self.spn_peso_nf.setStyleSheet(self._input_style())
        self.spn_peso_nf.valueChanged.connect(self._calcular_total_nota)
        return self.spn_peso_nf

    def _create_numero_nf(self):
        self.txt_numero_nf = QLineEdit()
        self.txt_numero_nf.setPlaceholderText("Número da Nota Fiscal")
        self.txt_numero_nf.setStyleSheet(self._input_style())
        return self.txt_numero_nf

    def _create_chave_nf(self):
        self.txt_chave_nf = QLineEdit()
        self.txt_chave_nf.setPlaceholderText("Chave de acesso da NF-e (44 dígitos)")
        self.txt_chave_nf.setStyleSheet(self._input_style())
        self.txt_chave_nf.setMaxLength(44)
        return self.txt_chave_nf

    def _create_valor_unitario(self):
        self.spn_valor_unitario = QDoubleSpinBox()
        self.spn_valor_unitario.setRange(0, 999999)
        self.spn_valor_unitario.setDecimals(4)
        self.spn_valor_unitario.setPrefix("R$ ")
        self.spn_valor_unitario.setStyleSheet(self._input_style())
        self.spn_valor_unitario.valueChanged.connect(self._calcular_total_nota)
        return self.spn_valor_unitario

    def _create_total_nota(self):
        self.spn_total_nota = QDoubleSpinBox()
        self.spn_total_nota.setRange(0, 999999999)
        self.spn_total_nota.setDecimals(2)
        self.spn_total_nota.setPrefix("R$ ")
        self.spn_total_nota.setStyleSheet(self._input_style())
        self.spn_total_nota.setReadOnly(True)
        return self.spn_total_nota

    def _calcular_peso_liquido(self):
        liquido = max(0, self.spn_peso_bruto.value() - self.spn_tara.value())
        self.spn_peso_liquido.setValue(liquido)

    def _calcular_total_nota(self):
        total = self.spn_peso_nf.value() * self.spn_valor_unitario.value()
        self.spn_total_nota.setValue(total)

    def _preencher(self, c):
        if c.get("data"):
            qdate = QDate.fromString(c["data"], "yyyy-MM-dd")
            if qdate.isValid():
                self.dt_data.setDate(qdate)

        idx = self.cmb_entidade.findData(c["entidade_id"])
        if idx >= 0:
            self.cmb_entidade.setCurrentIndex(idx)

        idx = self.cmb_lote.findData(c["lote_id"])
        if idx >= 0:
            self.cmb_lote.setCurrentIndex(idx)

        self.txt_placa.setText(c.get("placa", ""))
        self.spn_peso_bruto.setValue(float(c.get("peso_bruto", 0)))
        self.spn_tara.setValue(float(c.get("tara", 0)))
        self.spn_peso_liquido.setValue(float(c.get("peso_liquido", 0)))
        self.spn_peso_ticket.setValue(float(c.get("peso_ticket", 0)))
        self.spn_peso_nf.setValue(float(c.get("peso_nf", 0)))
        self.txt_numero_nf.setText(c.get("numero_nf", ""))
        self.txt_chave_nf.setText(c.get("chave_nf", ""))
        self.spn_valor_unitario.setValue(float(c.get("valor_unitario", 0)))
        self.spn_total_nota.setValue(float(c.get("total_nota", 0)))

    def _validar_salvar(self):
        if self.cmb_entidade.currentData() is None:
            self._msg_box(QMessageBox.Warning, "Validação", "Selecione uma entidade.")
            return
        if self.cmb_lote.currentData() is None:
            self._msg_box(QMessageBox.Warning, "Validação", "Selecione um lote.")
            return
        self.accept()

    def obter_dados(self):
        return {
            "data": self.dt_data.date().toString("yyyy-MM-dd"),
            "entidade_id": self.cmb_entidade.currentData(),
            "lote_id": self.cmb_lote.currentData(),
            "placa": self.txt_placa.text().strip().upper(),
            "peso_bruto": self.spn_peso_bruto.value(),
            "tara": self.spn_tara.value(),
            "peso_liquido": self.spn_peso_liquido.value(),
            "peso_ticket": self.spn_peso_ticket.value(),
            "peso_nf": self.spn_peso_nf.value(),
            "numero_nf": self.txt_numero_nf.text().strip(),
            "chave_nf": self.txt_chave_nf.text().strip(),
            "valor_unitario": self.spn_valor_unitario.value(),
            "total_nota": self.spn_total_nota.value(),
        }

    def _msg_box(self, icone, titulo, texto, botoes=None):
        return msg_box(self, icone, titulo, texto, botoes)

    def keyPressEvent(self, event):
        super().keyPressEvent(event)


class CarregamentoView(QWidget):
    def __init__(self):
        super().__init__()
        self.controller = CarregamentoController()
        self._setup_ui()


    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(15)

        header = QLabel("Carregamentos")
        header.setStyleSheet("color: white; font-size: 22px; font-weight: 700; letter-spacing: 2px;")
        layout.addWidget(header)

        card = QFrame()
        card.setStyleSheet("QFrame { background: white; border-radius: 20px; }")

        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(60)
        shadow.setColor(QColor(0, 0, 0, 30))
        shadow.setOffset(0, 8)
        card.setGraphicsEffect(shadow)

        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(20, 20, 20, 20)
        card_layout.setSpacing(12)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)

        self._lotes_filtro = self.controller.listar_lotes_carregamento(apenas_ativos=False)

        self.cmb_busca_lote = QComboBox()
        self.cmb_busca_lote.addItem("Todos os lotes", None)
        for l in self._lotes_filtro:
            nome = l["nome_lote"] + (" (inativo)" if not l["ativo"] else "")
            self.cmb_busca_lote.addItem(nome, l["id"])
        self.cmb_busca_lote.setStyleSheet(self._combo_style())
        self.cmb_busca_lote.currentIndexChanged.connect(self._carregar_dados)
        toolbar.addWidget(self.cmb_busca_lote)

        self.cmb_filtro_cliente = QComboBox()
        self.cmb_filtro_cliente.addItem("Todos os clientes", None)
        clientes = self.controller.listar_clientes_carregamento()
        for c in clientes:
            self.cmb_filtro_cliente.addItem(f"{c['razao_social']} ({c['cnpj_cpf']})", c["id"])
        self.cmb_filtro_cliente.setStyleSheet(self._combo_style())
        self.cmb_filtro_cliente.currentIndexChanged.connect(self._on_filtro_cliente_changed)
        toolbar.addWidget(self.cmb_filtro_cliente)

        lbl_de = QLabel("De:")
        lbl_de.setStyleSheet("color: #555; font-size: 11px; font-weight: 700; margin-left: 5px;")
        toolbar.addWidget(lbl_de)
        self.dt_inicio = QDateEdit()
        self.dt_inicio.setCalendarPopup(True)
        self.dt_inicio.setDisplayFormat("dd/MM/yyyy")
        self.dt_inicio.setSpecialValueText("Início")
        self.dt_inicio.setDate(self.dt_inicio.minimumDate())
        estilizar_calendario(self.dt_inicio)
        self.dt_inicio.setStyleSheet(self._date_style())
        self.dt_inicio.dateChanged.connect(self._carregar_dados)
        toolbar.addWidget(self.dt_inicio)

        lbl_a = QLabel("até")
        lbl_a.setStyleSheet("color: #4E342E; font-weight: 600; padding: 0 4px;")

        self.dt_fim = QDateEdit()
        self.dt_fim.setCalendarPopup(True)
        self.dt_fim.setDisplayFormat("dd/MM/yyyy")
        self.dt_fim.setSpecialValueText("Fim")
        self.dt_fim.setDate(self.dt_fim.minimumDate())
        estilizar_calendario(self.dt_fim)
        self.dt_fim.setStyleSheet(self._date_style())
        self.dt_fim.dateChanged.connect(self._carregar_dados)
        toolbar.addWidget(self.dt_fim)

        btn_atualizar = QPushButton("🔄 Atualizar")
        btn_atualizar.setCursor(Qt.PointingHandCursor)
        btn_atualizar.setFixedHeight(40)
        btn_atualizar.setStyleSheet("""
            QPushButton {
                padding: 8px 16px; background: #3498db; color: white;
                border: none; border-radius: 8px; font-weight: 700; font-size: 12px;
            }
            QPushButton:hover { background: #5dade2; }
            QPushButton:pressed { background: #2980b9; }
        """)
        btn_atualizar.clicked.connect(self._carregar_dados)
        toolbar.addWidget(btn_atualizar)

        toolbar.addStretch()

        self.btn_novo = QPushButton("+ Novo Carregamento")
        self.btn_novo.setCursor(Qt.PointingHandCursor)
        self.btn_novo.setFixedHeight(40)
        self.btn_novo.setStyleSheet(self._btn_primary())
        self.btn_novo.clicked.connect(self._novo)
        toolbar.addWidget(self.btn_novo)

        self.btn_editar = QPushButton("Editar")
        self.btn_editar.setCursor(Qt.PointingHandCursor)
        self.btn_editar.setFixedHeight(40)
        self.btn_editar.setStyleSheet(self._btn_warning())
        self.btn_editar.clicked.connect(self._editar)
        toolbar.addWidget(self.btn_editar)

        self.btn_excluir = QPushButton("Excluir")
        self.btn_excluir.setCursor(Qt.PointingHandCursor)
        self.btn_excluir.setFixedHeight(40)
        self.btn_excluir.setStyleSheet(self._btn_danger())
        self.btn_excluir.clicked.connect(self._excluir)
        toolbar.addWidget(self.btn_excluir)

        btn_exportar = QPushButton("📤 Exportar")
        btn_exportar.setCursor(Qt.PointingHandCursor)
        btn_exportar.setFixedHeight(40)
        btn_exportar.setStyleSheet("""
            QPushButton {
                padding: 8px 16px; background: #795548; color: white;
                border: none; border-radius: 8px; font-weight: 700; font-size: 12px;
            }
            QPushButton:hover { background: #8D6E63; }
            QPushButton:pressed { background: #5D4037; }
        """)
        btn_exportar.clicked.connect(self._exportar)
        toolbar.addWidget(btn_exportar)

        card_layout.addLayout(toolbar)

        self.table = QTableWidget()
        self.table.setColumnCount(14)
        self.table.setHorizontalHeaderLabels(
            ["ID", "Data", "Entidade", "Lote", "Placa", "P. Bruto", "Tara",
             "P. Líquido", "P. Ticket", "P. NF", "Nº NF", "Chave NF",
             "Valor Unit.", "Total NF"]
        )
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.verticalHeader().hide()
        self.table.doubleClicked.connect(self._editar)

        self._header = self.table.horizontalHeader()
        for i in range(14):
            self._header.setSectionResizeMode(i, QHeaderView.Interactive)
        self._header.setSectionResizeMode(2, QHeaderView.Stretch)
        self._header.setSectionResizeMode(3, QHeaderView.Stretch)

        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #e0e0e0;
                border-radius: 10px;
                background: #ffffff;
                gridline-color: transparent;
                font-size: 13px;
                color: #000;
                selection-background-color: #d4d4d4;
            }
            QTableWidget::item {
                padding: 8px 14px;
            }
            QTableWidget::item:selected {
                background: #d4d4d4;
                color: #1a1a1a;
            }
            QTableWidget::item:hover {
                background: #dcdcdc;
            }
            QHeaderView::section {
                background: #F5F0EB;
                color: #4E342E;
                padding: 10px 14px;
                border: none;
                border-bottom: 2px solid #795548;
                border-right: 1px solid #e8e8e8;
                font-weight: 700;
                font-size: 12px;
                letter-spacing: 0.5px;
            }
            QHeaderView::section:last {
                border-right: none;
            }
            QTableWidget::item:alternate {
                background: #e6e6e6;
            }
        """)

        card_layout.addWidget(self.table)

        self.lbl_resumo = QLabel()
        self.lbl_resumo.setStyleSheet("color: #4E342E; font-size: 14px; font-weight: 700; padding: 10px 0;")
        self.lbl_resumo.setAlignment(Qt.AlignRight)
        card_layout.addWidget(self.lbl_resumo)

        layout.addWidget(card)

    def _fmt_num(self, val):
        return f"{val:_.0f}".replace("_", ".")

    def _fmt_money(self, val):
        return f"{val:_.2f}".replace(".", ",").replace("_", ".")

    def _on_filtro_cliente_changed(self):
        entidade_id = self.cmb_filtro_cliente.currentData()
        self.cmb_busca_lote.blockSignals(True)
        self.cmb_busca_lote.clear()
        self.cmb_busca_lote.addItem("Todos os lotes", None)
        for l in self._lotes_filtro:
            if entidade_id is None or l["entidade_id"] == entidade_id:
                nome = l["nome_lote"] + (" (inativo)" if not l["ativo"] else "")
                self.cmb_busca_lote.addItem(nome, l["id"])
        self.cmb_busca_lote.blockSignals(False)
        self._carregar_dados()

    def _carregar_dados(self):
        entidade_id = self.cmb_filtro_cliente.currentData()
        lote_id = self.cmb_busca_lote.currentData()
        data_inicio = self.dt_inicio.date().toString("yyyy-MM-dd") if self.dt_inicio.text() != self.dt_inicio.specialValueText() else None
        data_fim = self.dt_fim.date().toString("yyyy-MM-dd") if self.dt_fim.text() != self.dt_fim.specialValueText() else None
        dados = self.controller.listar(entidade_id=entidade_id, lote_id=lote_id,
                                       data_inicio=data_inicio, data_fim=data_fim)
        self.table.setRowCount(len(dados))
        total_pesoTicket = total_pesoNF = total_nota = 0
        for i, c in enumerate(dados):
            data_raw = c.get("data", "")
            if data_raw:
                qd = QDate.fromString(data_raw, "yyyy-MM-dd")
                data_fmt = qd.toString("dd/MM/yyyy") if qd.isValid() else data_raw
            else:
                data_fmt = ""
            self.table.setItem(i, 0, QTableWidgetItem(str(c["id"])))
            self.table.setItem(i, 1, QTableWidgetItem(data_fmt))
            self.table.setItem(i, 2, QTableWidgetItem(c.get("entidade_nome", "")))
            self.table.setItem(i, 3, QTableWidgetItem(c.get("nome_lote", "")))
            self.table.setItem(i, 4, QTableWidgetItem(c.get("placa", "")))
            self.table.setItem(i, 5, QTableWidgetItem(self._fmt_num(c.get('peso_bruto', 0))))
            self.table.setItem(i, 6, QTableWidgetItem(self._fmt_num(c.get('tara', 0))))
            self.table.setItem(i, 7, QTableWidgetItem(self._fmt_num(c.get('peso_liquido', 0))))
            self.table.setItem(i, 8, QTableWidgetItem(self._fmt_num(c.get('peso_ticket', 0))))
            self.table.setItem(i, 9, QTableWidgetItem(self._fmt_num(c.get('peso_nf', 0))))
            self.table.setItem(i, 10, QTableWidgetItem(c.get("numero_nf", "")))
            self.table.setItem(i, 11, QTableWidgetItem(c.get("chave_nf", "")))
            self.table.setItem(i, 12, QTableWidgetItem(f"{c.get('valor_unitario', 0):.4f}"))
            self.table.setItem(i, 13, QTableWidgetItem(f"{c.get('total_nota', 0):.2f}"))
            total_pesoTicket += c.get("peso_ticket", 0) or 0
            total_pesoNF += c.get("peso_nf", 0) or 0
            total_nota += c.get("total_nota", 0) or 0
        self.table.resizeColumnsToContents()
        self._header.setSectionResizeMode(2, QHeaderView.Stretch)
        self._header.setSectionResizeMode(3, QHeaderView.Stretch)

        nome = self.cmb_filtro_cliente.currentText()
        self.lbl_resumo.setText(
            f"{nome}\n\n"
            f"    P. Líquido (Ticket):  {self._fmt_num(total_pesoTicket)} kg\n\n"
            f"    P. Líquido (NF):       {self._fmt_num(total_pesoNF)} kg\n\n"
            f"    Total NF:                 R$ {self._fmt_money(total_nota)}"
        )

    def _novo(self):
        dialog = CarregamentoDialog(self)
        if dialog.exec():
            dados = dialog.obter_dados()
            self.controller.salvar(dados)
            self._carregar_dados()

    def _editar(self):
        row = self.table.currentRow()
        if row < 0:
            self._msg_box(QMessageBox.Information, "Selecione", "Selecione um carregamento para editar.")
            return
        if not self._confirmar_senha():
            return
        c_id = int(self.table.item(row, 0).text())
        carregamento = self.controller.buscar_por_id(c_id)
        if not carregamento:
            self._msg_box(QMessageBox.Warning, "Erro", "Carregamento não encontrado.")
            return
        dialog = CarregamentoDialog(self, carregamento)
        if dialog.exec():
            dados = dialog.obter_dados()
            self.controller.atualizar(c_id, dados)
            self._carregar_dados()

    def _excluir(self):
        row = self.table.currentRow()
        if row < 0:
            self._msg_box(QMessageBox.Information, "Selecione", "Selecione um carregamento para excluir.")
            return
        if not self._confirmar_senha():
            return
        c_id = int(self.table.item(row, 0).text())
        confirm = self._msg_box(
            QMessageBox.Question, "Confirmar",
            "Tem certeza que deseja excluir este carregamento?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if confirm == QMessageBox.Yes:
            self.controller.remover(c_id)
            self._carregar_dados()

    def _exportar(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        path, _ = QFileDialog.getSaveFileName(
            self, "Exportar Relatório", "relatorio_carregamentos.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        entidade_id = self.cmb_filtro_cliente.currentData()
        lote_id = self.cmb_busca_lote.currentData()
        data_inicio = self.dt_inicio.date().toString("yyyy-MM-dd") if self.dt_inicio.text() != self.dt_inicio.specialValueText() else None
        data_fim = self.dt_fim.date().toString("yyyy-MM-dd") if self.dt_fim.text() != self.dt_fim.specialValueText() else None
        dados = self.controller.listar(entidade_id=entidade_id, lote_id=lote_id,
                                       data_inicio=data_inicio, data_fim=data_fim)

        wb = Workbook()
        ws = wb.active
        ws.title = "Carregamentos"

        cabecalhos = ["ID", "Data", "Entidade", "Lote", "Placa",
                       "P. Bruto", "Tara", "P. Líquido", "P. Ticket",
                       "P. NF", "Nº NF", "Chave NF", "Valor Unit.", "Total NF"]
        cols = len(cabecalhos)

        header_fill = PatternFill(start_color="795548", end_color="795548", fill_type="solid")
        header_font = Font(color="ffffff", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, titulo in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col, value=titulo)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for i, c in enumerate(dados, 2):
            data_raw = c.get("data", "")
            if data_raw:
                qd = QDate.fromString(data_raw, "yyyy-MM-dd")
                data_fmt = qd.toString("dd/MM/yyyy") if qd.isValid() else data_raw
            else:
                data_fmt = ""
            vals = [
                int(c["id"]), data_fmt, c.get("entidade_nome", ""), c.get("nome_lote", ""),
                c.get("placa", ""), c.get("peso_bruto", 0), c.get("tara", 0),
                c.get("peso_liquido", 0), c.get("peso_ticket", 0),
                c.get("peso_nf", 0), c.get("numero_nf", ""),
                c.get("chave_nf", ""),
                c.get("valor_unitario", 0), c.get("total_nota", 0),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                    if col in (1, 6, 7, 8, 9, 10):
                        cell.number_format = '#,##0'
                    elif col == 13:
                        cell.number_format = '#,##0.0000'
                    else:
                        cell.number_format = '#,##0.00'

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        for col_letra in ['F', 'G', 'H', 'I', 'J', 'L', 'M', 'N']:
            ws.column_dimensions[col_letra].width = 14
        ws.column_dimensions['K'].width = 18

        linha_total = len(dados) + 2
        ws.cell(row=linha_total, column=1).value = ""
        total_bruto = sum((c.get("peso_bruto", 0) or 0) for c in dados)
        total_liquido = sum((c.get("peso_liquido", 0) or 0) for c in dados)
        total_nf = sum((c.get("peso_nf", 0) or 0) for c in dados)
        total_nota = sum((c.get("total_nota", 0) or 0) for c in dados)
        ws.cell(row=linha_total, column=5, value="TOTAIS:").font = Font(bold=True)
        ws.cell(row=linha_total, column=6, value=total_bruto).number_format = '#,##0'
        ws.cell(row=linha_total, column=8, value=total_liquido).number_format = '#,##0'
        ws.cell(row=linha_total, column=10, value=total_nf).number_format = '#,##0'
        ws.cell(row=linha_total, column=14, value=total_nota).number_format = '#,##0.00'
        for col in range(1, cols + 1):
            ws.cell(row=linha_total, column=col).border = thin_border
            ws.cell(row=linha_total, column=col).font = Font(bold=True)

        wb.save(path)
        self._msg_box(QMessageBox.Information, "Exportado",
                       f"Relatório salvo em:\n{path}")

    def _msg_box(self, icone, titulo, texto, botoes=None):
        return msg_box(self, icone, titulo, texto, botoes)

    def _confirmar_senha(self):
        senha, ok = QInputDialog.getText(self, "Autenticação", "Digite sua senha:", QLineEdit.Password)
        if not ok or not senha:
            return False
        auth = AuthController()
        if auth.verificar_senha(senha):
            return True
        self._msg_box(QMessageBox.Warning, "Erro", "Senha incorreta.")
        return False

    def _btn_primary(self):
        return """
            QPushButton {
                padding: 8px 20px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5D4037, stop:1 #795548);
                color: white; border: none; border-radius: 8px;
                font-weight: 700; font-size: 12px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #795548, stop:1 #8D6E63);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3E2723, stop:1 #5D4037);
            }
        """

    def _btn_warning(self):
        return """
            QPushButton {
                padding: 8px 20px; background: #f39c12; color: white;
                border: none; border-radius: 8px; font-weight: 700; font-size: 12px;
            }
            QPushButton:hover { background: #f1c40f; }
            QPushButton:pressed { background: #d68910; }
        """

    def _btn_danger(self):
        return """
            QPushButton {
                padding: 8px 20px; background: #e74c3c; color: white;
                border: none; border-radius: 8px; font-weight: 700; font-size: 12px;
            }
            QPushButton:hover { background: #ec7063; }
            QPushButton:pressed { background: #c0392b; }
        """

    def _combo_style(self):
        return """
            QComboBox {
                padding: 10px 14px;
                border: 2px solid #ddd;
                border-radius: 10px;
                font-size: 13px;
                background: #fafafa;
                min-width: 200px;
                color: #000;
            }
            QComboBox:focus { border-color: #795548; background: white; }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 30px;
                border-left: 1px solid #ddd;
                border-top-right-radius: 10px;
                border-bottom-right-radius: 10px;
            }
            QComboBox QAbstractItemView {
                color: #000;
                background: white;
                selection-background-color: #e6e6e6;
                selection-color: #000;
            }
        """

    def _date_style(self):
        return """
            QDateEdit {
                padding: 8px 10px;
                border: 2px solid #ddd;
                border-radius: 8px;
                font-size: 12px;
                background: #fafafa;
                max-width: 130px;
                color: #000;
            }
            QDateEdit:focus { border-color: #795548; background: white; }
            QDateEdit::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 24px;
                border-left: 1px solid #ddd;
            }
        """
